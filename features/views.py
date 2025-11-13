from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.http import HttpResponse, JsonResponse, Http404
from django.views.decorators.http import require_http_methods
from django.core.exceptions import PermissionDenied
from django.conf import settings
from .models import Campaign, Donation, DonorProfile, Expense
from .forms import CampaignForm, DonationForm, DonorProfileForm, ExpenseForm, ContactForm
from django.db import models
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
import os
import mimetypes
from datetime import datetime
import uuid
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
logger = logging.getLogger(__name__)


def home(request):
    active_campaigns = Campaign.objects.filter(
        is_active=True, end_date__gte=timezone.now().date()
    ).order_by("-created_at")

    # Impact Stats
    total_funds_raised = (
        Donation.objects.filter(status="COMPLETED").aggregate(total=Sum("amount"))[
            "total"
        ]
        or 0
    )
    donor_count = (
        Donation.objects.filter(status="COMPLETED").values("donor").distinct().count()
    )
    projects_completed_count = Campaign.objects.filter(
        is_active=False, collected_amount__gte=models.F("target_amount")
    ).count()

    context = {
        "active_campaigns": active_campaigns,
        "total_funds_raised": total_funds_raised,
        "donor_count": donor_count,
        "projects_completed_count": projects_completed_count,
    }
    return render(request, "features/home.html", context)


def campaign_detail(request, campaign_id):
    campaign = get_object_or_404(Campaign, pk=campaign_id)
    donations = Donation.objects.filter(
        campaign=campaign, status="COMPLETED"
    ).order_by("-donation_date")[:10]

    # Check if the current user has donated to this campaign
    user_has_donated = False
    if request.user.is_authenticated:
        user_has_donated = Donation.objects.filter(
            campaign=campaign, 
            donor=request.user, 
            status="COMPLETED"
        ).exists()

    context = {
        "campaign": campaign,
        "donations": donations,
        "form": DonationForm() if request.user.is_authenticated else None,
        "user_has_donated": user_has_donated,
    }
    return render(request, "features/campaign_detail.html", context)


@login_required
def make_donation(request, campaign_id):
    campaign = get_object_or_404(Campaign, pk=campaign_id)

    if request.method == "POST":
        form = DonationForm(request.POST)
        if form.is_valid():
            donation = form.save(commit=False)
            donation.donor = request.user
            donation.campaign = campaign
            # Ensure initial status is pending until payment is confirmed
            donation.status = "PENDING"
            donation.save()

            # Treat as immediate confirmation for now
            donation.status = "COMPLETED"
            donation.transaction_id = f"OFFLINE-{donation.id}"
            donation.save()

            campaign.collected_amount += donation.amount
            campaign.save()

            profile, created = DonorProfile.objects.get_or_create(user=request.user)
            profile.total_donations += donation.amount
            profile.last_donation_date = timezone.now()
            profile.save()

            subject = "Thank you for your donation!"
            message = render_to_string(
                "features/email/thank_you.html",
                {"donation": donation, "campaign": campaign, "user": request.user},
            )
            send_mail(
                subject,
                "",
                "noreply@villagefunds.com",
                [request.user.email],
                html_message=message,
            )

            messages.success(request, "Thank you for your donation!")
            return redirect("features:campaign_detail", campaign_id=campaign.id)
    else:
        form = DonationForm()

    return render(
        request, "features/make_donation.html", {"form": form, "campaign": campaign}
    )




@login_required
def donor_profile(request):
    profile, created = DonorProfile.objects.get_or_create(user=request.user)
    donations = Donation.objects.filter(donor=request.user).order_by("-donation_date")
    donations_completed_count = Donation.objects.filter(
        donor=request.user, status="COMPLETED"
    ).count()

    # Compute profile completion percentage based on filled fields
    filled = 0
    total = 3  # phone_number, address, photo
    if getattr(profile, "phone_number", None):
        filled += 1
    if getattr(profile, "address", None):
        filled += 1
    if getattr(profile, "photo", None):
        filled += 1
    profile_completion = round((filled / total) * 100) if total else 0

    if request.method == "POST":
        form = DonorProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Profile updated successfully!")
            return redirect("features:donor_profile")
    else:
        form = DonorProfileForm(instance=profile)

    context = {
        "form": form,
        "profile": profile,
        "donations": donations,
        "donations_completed_count": donations_completed_count,
        "profile_completion": profile_completion,
    }
    return render(request, "features/donor_profile.html", context)


@login_required
def campaign_list(request):
    campaigns = Campaign.objects.all().order_by("-created_at")
    return render(request, "features/campaign_list.html", {"campaigns": campaigns})


@login_required
def add_campaign(request):
    if request.method == "POST":
        form = CampaignForm(request.POST, request.FILES)
        if form.is_valid():
            campaign = form.save()
            messages.success(request, "Campaign created successfully!")
            return redirect("features:campaign_detail", campaign_id=campaign.id)
    else:
        form = CampaignForm()

    return render(request, "features/add_campaign.html", {"form": form})


@login_required
def add_expense(request, campaign_id):
    campaign = get_object_or_404(Campaign, pk=campaign_id)

    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.campaign = campaign
            expense.approved_by = request.user
            expense.save()
            messages.success(request, "Expense added successfully!")
            return redirect("features:campaign_detail", campaign_id=campaign.id)
    else:
        form = ExpenseForm()

    return render(
        request, "features/add_expense.html", {"form": form, "campaign": campaign}
    )

def fund_usage(request):
    donations = Donation.objects.filter(status="COMPLETED").order_by("-donation_date")[
        :50
    ]
    expenses = Expense.objects.all().order_by("-date")[:50]
    return render(
        request,
        "features/fund_usage.html",
        {"donations": donations, "expenses": expenses},
    )


def about(request):
    return render(request, "features/about.html")


def gallery(request):
    # Placeholder: pass empty list for now
    gallery_items = []
    return render(request, "features/gallery.html", {"gallery_items": gallery_items})


def contact(request):
    # Math CAPTCHA setup
    if request.method == "GET":
        try:
            import random
            a = random.randint(1, 9)
            b = random.randint(1, 9)
            expected = a + b
            request.session["contact_captcha_a"] = a
            request.session["contact_captcha_b"] = b
            request.session["contact_captcha"] = expected
            request.session.modified = True
        except Exception:
            a, b, expected = 2, 3, 5
            request.session["contact_captcha_a"] = a
            request.session["contact_captcha_b"] = b
            request.session["contact_captcha"] = expected
            request.session.modified = True
        form = ContactForm(expected_captcha=expected)
        return render(
            request,
            "features/contact.html",
            {
                "form": form,
                "captcha_question": f"{a} + {b}",
                "instructions": "Fill in all required fields. Attach relevant files if needed. Solve the simple math CAPTCHA to verify you are human.",
            },
        )

    # POST: validate and deliver
    a = request.session.get("contact_captcha_a", 0)
    b = request.session.get("contact_captcha_b", 0)
    expected = request.session.get("contact_captcha", None)
    form = ContactForm(request.POST, request.FILES, expected_captcha=expected)

    if form.is_valid():
        data = form.cleaned_data
        # Determine admin recipient
        admin_recipient = getattr(settings, "ADMIN_CONTACT_EMAIL", None) or getattr(settings, "EMAIL_HOST_USER", None) or getattr(settings, "DEFAULT_FROM_EMAIL", None)
        from_email = getattr(settings, "DEFAULT_FROM_EMAIL", getattr(settings, "EMAIL_HOST_USER", None))
        if not admin_recipient:
            messages.error(request, "Admin contact email is not configured.")
            return render(request, "features/contact.html", {"form": form, "instructions": instructions, "captcha_question": captcha_question})
        # Compose admin email
        ticket_id = uuid.uuid4().hex[:8].upper()
        admin_subject = f"[Contact #{ticket_id}] {data.get('subject') or 'No subject'}"
        if data.get("category"):
            admin_subject = f"[{data['category']}] {admin_subject}"
        admin_body = (
            f"New contact submission on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
            f"Ticket ID: {ticket_id}\n"
            f"Name: {data['name']}\n"
            f"Email: {data['email']}\n"
            f"Category: {data.get('category') or 'Not specified'}\n"
            f"Message:\n{data['message']}\n"
        )
        admin_email = EmailMessage(
            admin_subject,
            admin_body,
            from_email,
            [admin_recipient],
            reply_to=[data["email"]],
        )
        # Robust attachment handling
        attachment = data.get("attachment")
        if attachment:
            try:
                try:
                    attachment.seek(0)
                except Exception:
                    pass
                filename = os.path.basename(getattr(attachment, "name", "attachment"))
                guessed = mimetypes.guess_type(filename)[0]
                ct = getattr(attachment, "content_type", None)
                content_type = ct if (ct and ct != "application/octet-stream") else (guessed or "application/pdf" if filename.lower().endswith(".pdf") else "application/octet-stream")
                # Prefer file-path attach when available (TemporaryUploadedFile)
                if hasattr(attachment, "temporary_file_path"):
                    admin_email.attach_file(attachment.temporary_file_path(), mimetype=content_type)
                else:
                    content = attachment.read()
                    admin_email.attach(filename, content, content_type)
                logger.info(
                    "Attached file to contact email: name=%s size=%s content_type=%s guessed=%s type=%s",
                    filename,
                    getattr(attachment, "size", "unknown"),
                    content_type,
                    guessed,
                    type(attachment).__name__,
                )
            except Exception as e:
                logger.error("Failed to attach file to admin email: %s", e, exc_info=True)
                messages.warning(request, "Attachment couldn't be added. Your message was sent without the file.")
        # Send admin email
        admin_email.send(fail_silently=False)
        # Send acknowledgment to user
        ack_subject = f"We received your message (Ticket {ticket_id})"
        ack_body = (
            f"Hello {data['name']},\n\n"
            f"Thanks for reaching out. We've received your message and will respond soon.\n"
            f"Ticket ID: {ticket_id}\n"
            f"Subject: {data.get('subject') or 'No subject'}\n"
            f"Category: {data.get('category') or 'Not specified'}\n\n"
            f"Your message:\n{data['message']}\n\n"
            f"Best regards,\nSupport Team"
        )
        ack_email = EmailMessage(ack_subject, ack_body, from_email, [data["email"]])
        ack_email.send(fail_silently=True)
        # Clear captcha session
        request.session.pop("contact_captcha_a", None)
        request.session.pop("contact_captcha_b", None)
        request.session.pop("contact_captcha", None)
        messages.success(request, "Your message has been sent. We've emailed a confirmation.")
        return redirect("features:contact")

    # Invalid form: re-render with errors
    return render(
        request,
        "features/contact.html",
        {"form": form, "captcha_question": f"{a} + {b}", "instructions": "Please correct the highlighted fields."},
    )


def faq(request):
    return render(request, "features/faq.html")


@login_required
def donation_list(request):
    """Display a list of donations for the current user"""
    donations = Donation.objects.filter(donor=request.user).order_by("-donation_date")
    return render(request, "features/donation_list.html", {"donations": donations})


@login_required
def donor_profile_list(request):
    """Display a list of all donor profiles (admin view)"""
    if not request.user.is_staff:
        messages.error(request, "You don't have permission to view this page.")
        return redirect("features:home")
    
    profiles = DonorProfile.objects.all().order_by("-total_donations")
    return render(request, "features/donor_profile_list.html", {"profiles": profiles})


def get_client_ip(request):
    """Get the client's IP address from the request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


@login_required
def download_campaign_donations(request, campaign_id):
    """Download all donations for a specific campaign as an Excel file"""
    campaign = get_object_or_404(Campaign, id=campaign_id)
    
    # Get all completed donations for this campaign
    donations = Donation.objects.filter(
        campaign=campaign,
        status='COMPLETED'
    ).select_related('donor').order_by('-donation_date')
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{campaign.title} - Donations"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Campaign info section
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Donations Report - {campaign.title}"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A3:H3')
    ws['A3'] = f"Total Donations: {donations.count()} | Total Amount: ₹{donations.aggregate(total=Sum('amount'))['total'] or 0:.2f}"
    ws['A3'].alignment = Alignment(horizontal="center")
    ws['A3'].font = Font(bold=True)
    
    # Empty row
    ws.append([])
    
    # Headers
    headers = [
        'Date', 'Donor Name', 'Email', 'Amount (₹)', 'Payment Method', 
        'Transaction ID', 'Status', 'Message'
    ]
    
    header_row = ws.max_row + 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for donation in donations:
        donor_name = "Anonymous"
        donor_email = ""
        
        if donation.donor:
            donor_name = f"{donation.donor.first_name} {donation.donor.last_name}".strip() or donation.donor.username
            donor_email = donation.donor.email
        
        row_data = [
            donation.donation_date.strftime('%Y-%m-%d %H:%M:%S'),
            donor_name,
            donor_email,
            float(donation.amount),
            donation.payment_method,
            donation.transaction_id or '',
            donation.status,
            donation.message or ''
        ]
        ws.append(row_data)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Format amount column as INR currency
    amount_col = 4  # Column D
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=amount_col)
        cell.number_format = '₹#,##0.00'
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    filename = f"{campaign.title.replace(' ', '_')}_donations_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        excel_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
